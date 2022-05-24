import os
import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors

# PROGRAM TO CHANGE MATERIAL PRICES IN EACH TAKEOFF FILE WITHIN A FOLDER
# Declerations
directory = '/home/abramnaranjo/.bkp/knells/Corridor'
folder = '/home/abramnaranjo/.bkp/knells/new-prices'
# Yellow Fill to note which cells/prices have been changed when spreadsheet is open
yellowFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='f6fa28'))


# Iterate through each file in the folder
for filename in os.listdir(directory):
    path = os.path.join(directory, filename)
    save = os.path.join(folder, filename)
    # VB must be set to true
    wb = openpyxl.load_workbook(path, read_only=False, keep_vba=True)
    # Makes the hidden sheet 'KAMCO' active
    sh = wb['KAMCO']
    sh.sheet_state = 'hidden'

    # Changing Prices within the loop
    #5/8 FIRECODE C89
    sh.cell(row=89, column=3).value = "0.516"
    sh.cell(row=89, column=3).fill = yellowFill


    #7/8 HAT CHANNEL 20GA Q56
    sh.cell(row=56, column=17).value = "0.963"
    sh.cell(row=56, column=17).fill = yellowFill

    #5/8 MOLDTOUGH/XP C93
    sh.cell(row=93, column=3).value = "0.636"
    sh.cell(row=93, column=3).fill = yellowFill

    #Mud B138-B146
    sh.cell(row=138, column=2).value = "21.45"
    sh.cell(row=138, column=2).fill = yellowFill
    sh.cell(row=139, column=2).value = "21.45"
    sh.cell(row=139, column=2).fill = yellowFill
    sh.cell(row=140, column=2).value = "21.45"
    sh.cell(row=140, column=2).fill = yellowFill
    sh.cell(row=141, column=2).value = "21.45"
    sh.cell(row=141, column=2).fill = yellowFill
    sh.cell(row=142, column=2).value = "21.45"
    sh.cell(row=142, column=2).fill = yellowFill
    sh.cell(row=143, column=2).value = "21.45"
    sh.cell(row=143, column=2).fill = yellowFill
    sh.cell(row=144, column=2).value = "21.45"
    sh.cell(row=144, column=2).fill = yellowFill
    sh.cell(row=145, column=2).value = "21.45"
    sh.cell(row=145, column=2).fill = yellowFill
    sh.cell(row=146, column=2).value = "21.45"
    sh.cell(row=146, column=2).fill = yellowFill

    #Tape B150
    sh.cell(row=150, column=2).value = "4.125"
    sh.cell(row=150, column=2).fill = yellowFill

    # 81/2" PAN HEAD FRAMING SCREWS E121
    sh.cell(row=121, column=5).value = "164.75"
    sh.cell(row=121, column=5).fill = yellowFill

    # 6X1 1/4" Fine Thread Screws
    sh.cell(row=122, column=2).value = "79.15"
    sh.cell(row=127, column=2).value = "79.15"
    sh.cell(row=122, column=2).fill = yellowFill
    sh.cell(row=127, column=2).fill = yellowFill

    #POWERS C5 3/4" Pin w/ Fuel 800/box
    sh.cell(row=121, column=12).value = "156.72"
    sh.cell(row=128, column=10).value = "156.72"
    sh.cell(row=121, column=12).fill = yellowFill
    sh.cell(row=128, column=10).fill = yellowFill

    #RC-1 25GA 1 1/2"
    sh.cell(row=56, column=16).value = "0.4"
    sh.cell(row=56, column=16).fill = yellowFill
    #20GA 1 1/2"
    sh.cell(row=57, column=17).value = "0.4"
    sh.cell(row=57, column=17).fill = yellowFill

    #1/4" High Flex GWB
    sh.cell(row=78, column=3).value = "0.677"
    sh.cell(row=78, column=3).fill = yellowFill

    wb.save(save)
